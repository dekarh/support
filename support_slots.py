# -*- coding: utf-8 -*-

from collections import OrderedDict 

#from datetime import datetime, timedelta, time, date
import datetime
import openpyxl
import os
import string
from subprocess import Popen, PIPE

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QTableWidgetItem, QAbstractItemView, QListWidgetItem

import sqlite3
from sqlite3 import Error

#from mysql.connector import MySQLConnection

from support_win import Ui_Form

from lib import read_config, l, s, fine_phone, format_phone, fine_snils

STATUSES = [ 'ПОКАЗ', 'берут', 'лучшиеЦена', 'лучшиеУсловия', '---', 'недозвон', 'согласовывают', '=не-там=', 'далеко',
            'НЕинтересно', 'ремонт-', 'мебель-', 'санузел-', 'НЕберут', 'НЕсегодня', 'НЕскоро',
            'НЕТфото','НЕадекват', 'ДУБЛЬ', 'СДАЛИ', 'КОРОТКИЙ', 'НЕИЗВ.СТАТУС']
STATUSES_ADD = ['ЕСЛИподнять', 'СамоПросмотр', 'НЕрегистрируют', 'дорого', 'комиссия-', 'ВИРТ']
CUTS = ('пгт', 'поселок городского типа',  'посёлок городского типа', 'пос', 'поселение', 'поселок', 'посёлок',
         'п', 'рп', 'рабочий посёлок', 'рабочий поселок', 'кп', 'курортный посёлок', 'курортный поселок', 'пс',
         'сс', 'смн', 'дп', 'дачный поселок', 'дачный посёлок', 'садовое товарищество',
         'садоводческое некоммерческое товарищество', 'садоводческое товарищество', 'снт', 'нп', 'пст', 'ж/д_ст',
         'ж/д ст', 'железнодорожная станция', 'с', 'село', 'м', 'д', 'дер', 'деревня', 'сл', 'ст', 'ст-ца',
         'станица', 'х', 'хут', 'хутор', 'рзд', 'у', 'урочище', 'клх', 'колхоз', 'свх', 'совхоз', 'зим', 'зимовье',
         'микрорайон', 'мкр', 'аллея', 'а', 'бульвар', 'б-р', 'бул', 'в/ч', 'военная часть', 'военный городок',
          'гск', 'гаражно-строительный кооператив', 'гк', 'гаражный кооператив', 'кв-л', 'квартал', 'линия',
         'лин', 'наб', 'набережная', 'переулок', 'пер', 'переезд', 'пл', 'площадь', 'пр-кт', 'пр-т', 'проспект', 'пр',
         'проезд', 'тер', 'терр', 'территория', 'туп', 'тупик', 'ул', 'улица', 'ш', 'шоссе')
STOPWORDS = ('мкр', 'микрорайон', 'область', 'обл', 'район', 'р-н' , 'поселок', 'посёлок', 'город', 'городок',
             'москва', 'московская', 'люберцы', 'реутов', 'апрелевка', 'балашиха', 'бронницы', 'верея', 'видное',
             'Волоколамск', 'Воскресенск', 'Высоковск', 'Голицыно', 'Дедовск', 'Дзержинский', 'Дмитров',
             'долгопрудный', 'домодедово', 'дрезна', 'дубна', 'егорьевск', 'жуковский', 'зарайск', 'звенигород',
             'ивантеевка', 'истра', 'кашира', 'клин', 'коломна', 'королёв', 'котельники', 'красноармейск',
             'красногорск', 'краснозаводск', 'краснознаменск', 'кубинка', 'куровское', 'ликино-дулёво', 'лобня',
             'лосино-петровский', 'луховицы', 'лыткарино', 'люберцы', 'можайск', 'мытищи', 'наро-фоминск', 'ногинск',
             'одинцово', 'озёры', 'орехово-Зуево', 'павловский посад', 'пересвет', 'подольск', 'протвино',
             'пушкино', 'пущино', 'раменское', 'реутов', 'рошаль', 'руза', 'сергиев посад', 'серпухов',
             'солнечногорск', 'старая купавна', 'ступино', 'талдом', 'фрязино', 'химки', 'хотьково',
             'черноголовка', 'чехов', 'шатура', 'щёлково', 'электрогорск', 'электросталь', 'электроугли', 'яхрома')

class MainWindowSlots(Ui_Form):   # Определяем функции, которые будем вызывать в слотах

    def setupUi(self, form):
        Ui_Form.setupUi(self,form)
        self.dateTimeEdit.hide()
        self.table2base = {}
        self.avitoIDs = tuple()
        self.cianIDs = tuple()
        self.lwStatuses.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.cutCostMin = l(self.leCostMin.text())
        self.cutCostMax = l(self.leCostMax.text())
        self.cutMetroMetersMax = l(self.leMetroMetersMax.text())
        self.cutMetroMinutesMax = l(self.leMetroMinutesMax.text())
        for status in STATUSES:
            item = QListWidgetItem(status)
            self.lwStatuses.addItem(item)
        self.items2cards = []
        clients = os.listdir('./clients')
        self.cmbFolders.addItems(clients)
        self.cmbFolders.setCurrentText(clients[0])
        self.cmbStatus.addItems(STATUSES)
        self.cutComission = int(self.leComission.text())
        self.cutStatuses = tuple(STATUSES)
        self.cardStatuses = []
        calls = []
        calls_in = os.listdir('incoming')
        for call in calls_in:
            calls.append('incoming/' + call)
        calls_ogg = [x for x in calls if x.endswith('.ogg')]
        calls_amr = [x for x in calls if x.endswith('.amr')]
        calls_wav = [x for x in calls if x.endswith('.wav')]
        calls_mp3 = [x for x in calls if x.endswith('.mp4')]
        self.calls = calls_ogg + calls_amr + calls_mp3 + calls_wav
        self.calls_ids = []
        self.changeDirectory()


        q1 = """
        self.client_id = None
        self.hasFileFolder = False
        self.dbconfig_crm = read_config(filename='alone.ini', section='crm')
        self.dbconfig_alone = read_config(filename='alone.ini', section='alone')
        self.alone_files = {}
        with open("all_files.txt", "rt") as file_all:
            for i, line in enumerate(file_all):
                if i > 1:
                    if len(line.split('/')) > 2 and line.find('search') == -1:
                        file_name = line.split('.wav')[0].split('/')[2].lower()
                        path_name = line.split('./recup_dir.')[1].split('/')[0].replace('/n','')
                        if self.alone_files.get(file_name, None):
                            self.alone_files[file_name].append(path_name)
                        else:
                            self.alone_files[file_name] = [path_name]
        self.twRezkeyPressEventMain = self.twRez.keyPressEvent
        self.twRez.keyPressEvent = self.twRezkeyPressEvent
        self.clbSave.setEnabled(False)
        self.contracts = {None:None}
        self.clbReport2xlsx.setEnabled(False)
        self.threads = []
        self.progressBar.hide()
        """
        return

    def changeDirectory(self):
        client = self.cmbFolders.currentText()
        self.cmbFolders.clear()
        self.cmbFolders.addItems(os.listdir('./clients'))
        self.cmbFolders.setCurrentText(client)
        self.con = sqlite3.connect('./clients/' + client + '/support.db',
                                   detect_types=sqlite3.PARSE_DECLTYPES|sqlite3.PARSE_COLNAMES)
        cursorObj = self.con.cursor()
        cursorObj.execute("SELECT count(*) FROM sqlite_master WHERE type='table' AND name='cards'")
        rows = cursorObj.fetchall()
        if not rows[0][0]:
            # создаем структуру БД
            cursorObj.execute("CREATE TABLE cards(id integer NOT NULL PRIMARY KEY, finderType integer, "
                              "idINfinder integer KEY, linkINfinder text, address text KEY, roomCount integer, "
                              "typeObj text, metro text, metroMeters integer, metroMinutes integer, square float, "
                              "floor integer, maxFloor integer, parking text, phone1 biginteger, phone2 biginteger, "
                              "phone3 biginteger, about text, remont text, roomSquare text, balcon text, windows text, "
                              "sanuzel text, withChildrensPets text, additional text, buildingSeria text, height float,"
                              " lift text, chute text, idObject integer, cardStatus integer, agentComission integer, "
                              "buyerComission integer, buildingType text, cost integer, sostav text, "
                              "status text, showingAT timestamp, zalog integer, squareLive float, "
                              "squareKitchen float, note text)")
            self.con.commit()
        self.loadBase()
        return

    def loadBase(self):
        cursorObj = self.con.cursor()
        cursorObj.execute("SELECT * FROM cards GROUP BY idINfinder ORDER BY address")
        self.cards = cursorObj.fetchall()
        self.colNames = {}
        for i, colName in enumerate(cursorObj.description):
            self.colNames[colName[0]] = i
        self.load_lwCards()
        self.click_lwCards()
        return

    def load_lwCards(self):
        cianIDs = []
        avitoIDs = []
        idTable = 0
        self.lwCards.clear()
        self.cardStatuses = []
        for k, card in enumerate(self.cards):
            self.cardStatuses.append('')
            if card[self.colNames['finderType']] == 1:
                cianIDs.append(int(card[self.colNames['idINfinder']]))
            if card[self.colNames['finderType']] == 2:
                avitoIDs.append(int(card[self.colNames['idINfinder']]))
            if not card[self.colNames['status']]:
                self.cardStatuses[k] = '---'
            elif card[self.colNames['status']] not in STATUSES:
                self.cardStatuses[k] = 'НЕИЗВ.СТАТУС'
            else:
                self.cardStatuses[k] = card[self.colNames['status']]
            if card[self.colNames['agentComission']] > self.cutComission:
                continue
            if self.cardStatuses[k] not in self.cutStatuses:
                continue
            if card[self.colNames['cost']] > self.cutCostMax:
                continue
            if card[self.colNames['cost']] < self.cutCostMin:
                continue
            if self.cutMetroMetersMax and card[self.colNames['metroMeters']]:
                if card[self.colNames['metroMeters']] > self.cutMetroMetersMax:
                    continue
            if self.cutMetroMinutesMax and card[self.colNames['metroMinutes']]:
                if card[self.colNames['metroMinutes']] > self.cutMetroMinutesMax:
                    continue
            self.table2base[idTable] = k
            idTable += 1
            addressList = str(card[self.colNames['address']]).lower().replace(',','').replace('.','').replace('  ',' ')\
                .replace('  ',' ').split(' ')
            address = ''
            addrs = []
            for adr in addressList:
                if adr not in CUTS:
                    addrs.append(adr)
            missed = -1
            if len(addrs[0]):
                for i, addr in enumerate(addrs):
                    if i == 0 and addr[0] in string.digits:
                        missed = i
                    elif missed > -1 and addr[0] in string.digits:
                        address += addrs[missed] + ' ' + addr + ' '
                        missed = -1
                    else:
                        address += addr + ' '
            if card[self.colNames['square']]:
                address += str(card[self.colNames['square']]) + 'м²'
            if card[self.colNames['roomCount']] != None:
                if card[self.colNames['roomCount']] == 0:
                    address += 'Студ'
                else:
                    address += str(card[self.colNames['roomCount']]) + 'к'
            if card[self.colNames['floor']] and card[self.colNames['maxFloor']]:
                address += str(card[self.colNames['floor']]) + '/' + str(card[self.colNames['maxFloor']]) + 'эт'
            item = QListWidgetItem(address.strip())
            self.lwCards.addItem(item)
        self.cianIDs = tuple(cianIDs)
        self.avitoIDs = tuple(avitoIDs)
        self.lblCount.setText(str(self.lwCards.count()))

    def click_lwCards(self):
        if self.lwCards.count():
            tableNum = 0
            if self.lwCards.currentIndex().row() > 0:
                tableNum = self.lwCards.currentIndex().row()
            cardNum = self.table2base[tableNum]
            self.lwCalls.clear()
            calls = []
            calls_in = os.listdir('incoming')
            for call in calls_in:
                calls.append('incoming/' + call)
            calls_ogg = [x for x in calls if x.endswith('.ogg')]
            calls_amr = [x for x in calls if x.endswith('.amr')]
            calls_wav = [x for x in calls if x.endswith('.wav')]
            calls_mp3 = [x for x in calls if x.endswith('.mp4')]
            self.calls = calls_ogg + calls_amr + calls_mp3 + calls_wav
            self.calls_ids = []
            for i, call in enumerate(self.calls):
                for phone in [self.cards[cardNum][self.colNames['phone1']],
                              self.cards[cardNum][self.colNames['phone2']],
                              self.cards[cardNum][self.colNames['phone3']]]:
                    if l(phone) > 70000000000:
                        if format_phone(call.split(']_[')[1]) == format_phone(phone):
                            self.calls_ids.append(i)
            cs = {}
            for call_id in self.calls_ids:
                a = self.calls[call_id]
                t = datetime.datetime(l(a.split(']_[')[2][6:]), l(a.split(']_[')[2][3:5]), l(a.split(']_[')[2][:2]),
                             l(a.split(']_[')[3][:2]), l(a.split(']_[')[3][3:5]), l(a.split(']_[')[3][6:8]))
                cs[t] = call_id
            calls_ids_buff = []
            for kk, i in sorted(cs.items(), key=lambda item: item[0]):  # Хитровычурная сортирвка с исп. sorted()
                calls_ids_buff.append(i)
            self.calls_ids = calls_ids_buff
            for i, call_id in enumerate(self.calls_ids):
                a = self.calls[call_id]
                t = datetime.datetime(l(a.split(']_[')[2][6:]), l(a.split(']_[')[2][3:5]), l(a.split(']_[')[2][:2]),
                             l(a.split(']_[')[3][:2]), l(a.split(']_[')[3][3:5]), l(a.split(']_[')[3][6:8]))
                self.lwCalls.addItem(QListWidgetItem(t.strftime('%d.%m.%y %H:%M') + ' ' +
                                                     fine_phone(l(a.split(']_[')[1]))))
            self.lePhone1.setText(fine_phone(self.cards[cardNum][self.colNames['phone1']]))
            self.lePhone2.setText(fine_phone(self.cards[cardNum][self.colNames['phone2']]))
            self.leNote.setText(self.cards[cardNum][self.colNames['note']])
            self.lblComission.setText(str(self.cards[cardNum][self.colNames['agentComission']]))
            self.cmbStatus.setCurrentText(self.cardStatuses[cardNum])
            if self.cards[cardNum][self.colNames['showingAT']]:
                self.dateTimeEdit.setDateTime(self.cards[cardNum][self.colNames['showingAT']])
            else:
                self.dateTimeEdit.setDateTime(datetime.datetime.now())

    def click_lwCalls(self, index=None):
        audios = ''
        for i, call_id in enumerate(self.calls_ids):
            audios +=  self.calls[call_id] + ' '
        proc = Popen('gnome-mpv ' + audios, shell=True, stdout=PIPE, stderr=PIPE)
        proc.wait()  # дождаться выполнения
        res = proc.communicate()  # получить tuple('stdout', 'stderr')
        if proc.returncode:
            print(res[1])
            print('result:', res[0])

    def click_clbImport(self):
        files = os.listdir('./clients/' + self.cmbFolders.currentText())
        for file in files:
            if file.startswith('offer') and file.endswith('.xlsx'):
                # Циан
                wb = openpyxl.load_workbook(filename='./clients/' + self.cmbFolders.currentText() + '/' + file,
                                            read_only=True)
                ws = wb[wb.sheetnames[0]]
                xlsx_header = {}
                for i, row in enumerate(ws):
                    if not i:
                        for j, cell in enumerate(row):
                            xlsx_header[cell.value] = j
                finderType = 1
                linkINfinder,address,typeObj,metro,parking,about,remont,roomSquare = '','','','','','','',''
                sanuzel,withChildrensPets,additional,lift,chute,buildingType,balcon,windows = '','','','','','','',''
                height,square,squareLive,squareKitchen = 0.0,0.0,0.0,0.0
                roomCount,floor,maxFloor,cost,zalog,agentComission,buyerComission = 0,0,0,0,0,0,0
                phone1, phone2, phone3, metroMinutes, idINfinder = 0, 0, 0, 0, 0
                for i, row in enumerate(ws):
                    if i:
                        for j, cell in enumerate(row):
                            if j == xlsx_header.get('ID  объявления', -1):
                                idINfinder = int(cell.value)
                                if not idINfinder:
                                    continue
                            elif j == xlsx_header.get('Количество комнат', -1):
                                roomCount = l(str(cell.value).split(',')[0])
                                # !!!!!!!!!!!!!!!!!!!!!! Недоделал Изолированные-неизолированные
                            elif j == xlsx_header.get('Тип', -1):
                                typeObj = cell.value
                            elif j == xlsx_header.get('Метро', -1):
                                metro = cell.value
                                metroMinutes = l(cell.value)
                            elif j == xlsx_header.get('Адрес', -1):
                                if str(cell.value).strip().strip('\n').strip().strip():
                                    addressStopped = ''
                                    for adr in str(cell.value).lower().split(','):
                                        stopped = False
                                        for stopword in STOPWORDS:
                                            if adr.find(stopword) > -1:
                                                stopped = True
                                        if not stopped:
                                            addressStopped += ' ' + adr
                                    addressList = addressStopped.strip().replace(',', '').replace('.','')\
                                        .replace('  ', ' ').replace('  ', ' ').split(' ')
                                    address = ''
                                    addrs = []
                                    for adr in addressList:
                                        if adr not in CUTS:
                                            addrs.append(adr)
                                    missed = -1
                                    if len(addrs[0]):
                                        for i, addr in enumerate(addrs):
                                            if i == 0 and addr[0] in string.digits:
                                                missed = i
                                            elif missed > -1 and addr[0] in string.digits:
                                                address += addrs[missed] + ' ' + addr + ' '
                                                missed = -1
                                            else:
                                                address += addr + ' '
                                    address = address.strip()
                                else:
                                    address = ''
                            elif j == xlsx_header.get('Площадь, м2', -1):
                                parts = str(cell.value).split('/')
                                for k, part in enumerate(parts):
                                    if k == 0:
                                        square = float(part)
                                    elif k == 1:
                                        squareLive = float(part)
                                    elif k == 2:
                                        squareKitchen = float(part)
                            elif j == xlsx_header.get('Дом', -1):
                                homes = str(cell.value).split(',')
                                for home in homes:
                                    if home.find('/') > -1:
                                        floor = l(home.split('/')[0])
                                        maxFloor = l(home.split('/')[1])
                                    else:
                                        buildingType = home
                            elif j == xlsx_header.get('Парковка', -1):
                                parking = cell.value
                            elif j == xlsx_header.get('Цена', -1):
                                parts = str(cell.value).split(',')
                                for k, part in enumerate(parts):
                                    if k == 0:
                                        cost = int(float(part.split(' руб')[0]))
                                    elif part.find('Залог') > -1:
                                        zalog = int(part.split('алог - ')[1].split(' руб')[0])
                                # !!!!!! Не доделал комунальные услуги, срок и предоплату !!!!!!!!!!!
                            elif j == xlsx_header.get('Комиссия', -1):
                                parts = str(cell.value).split(',')
                                if len(parts) > 1:
                                    for part in parts:
                                        if part.find('кл') > -1:
                                            buyerComission = int(l(part))
                                        elif part.find('аг') > -1:
                                            agentComission = int(l(part))
                                else:
                                    if parts[0].strip() != '':
                                        buyerComission, agentComission = l(parts[0]),l(parts[0])
                            elif j == xlsx_header.get('Телефоны', -1):
                                parts = str(cell.value).split(',')
                                if len(parts) > 0:
                                    phone1 = format_phone(parts[0])
                                if len(parts) > 1:
                                    phone2 = format_phone(parts[1])
                                if len(parts) > 2:
                                    phone3 = format_phone(parts[2])
                            elif j == xlsx_header.get('Описание', -1):
                                about = cell.value
                            elif j == xlsx_header.get('Ремонт', -1):
                                remont = cell.value
                            elif j == xlsx_header.get('Площадь комнат, м2', -1):
                                roomSquare = cell.value
                            elif j == xlsx_header.get('Балкон', -1):
                                balcon = cell.value
                            elif j == xlsx_header.get('Окна', -1):
                                windows = cell.value
                            elif j == xlsx_header.get('Санузел', -1):
                                sanuzel = cell.value
                            elif j == xlsx_header.get('Можно с детьми/животными', -1):
                                withChildrensPets = cell.value
                            elif j == xlsx_header.get('Дополнительно', -1):
                                additional = cell.value
                            elif j == xlsx_header.get('Лифт', -1):
                                lift = cell.value
                            elif j == xlsx_header.get('Мусоропровод', -1):
                                chute = cell.value
                            elif j == xlsx_header.get('Ссылка на объявление', -1):
                                linkINfinder = cell.value
                            elif j == xlsx_header.get('Высота потолков, м', -1):
                                if cell.value:
                                    height = float(cell.value)
                        cursorObj = self.con.cursor()
                        if idINfinder:
                            if idINfinder in self.cianIDs:
                                cursorObj.execute("""UPDATE cards SET linkINfinder = ?, address = ?, roomCount = ?,
                                typeObj = ?, metro = ?, square = ?, floor = ?, maxFloor = ?, parking = ?, 
                                phone1 = ?, phone2 = ?, phone3 = ?, about = ?, remont = ?, roomSquare = ?,
                                balcon = ?, windows = ?, sanuzel = ?, withChildrensPets = ?, additional = ?,
                                height = ?, lift = ?, chute = ?, agentComission = ?, buyerComission = ?,
                                buildingType = ?, cost = ?, zalog = ?, squareLive = ?, squareKitchen = ?,
                                metroMinutes = ? 
                                WHERE idINfinder = ?""",
                                (linkINfinder, address,
                                roomCount, typeObj, metro, square, floor, maxFloor, parking, phone1, phone2, phone3,
                                about, remont, roomSquare, balcon, windows, sanuzel, withChildrensPets, additional,
                                height, lift, chute, agentComission, buyerComission, buildingType, cost, zalog,
                                squareLive,squareKitchen,metroMinutes,idINfinder))
                            else:
                                cursorObj.execute("INSERT INTO cards (finderType,idINfinder,linkINfinder,address,"
                                                  "roomCount,typeObj,metro,square,floor,maxFloor,parking,phone1,phone2,"
                                                  "phone3,about,remont,roomSquare,balcon,windows,sanuzel,"
                                                  "withChildrensPets,additional,height,lift,chute,agentComission,"
                                                  "buyerComission,buildingType,cost,zalog,squareLive,squareKitchen,"
                                                  "metroMinutes) "
                                                  "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?"
                                                  ",?,?)", (finderType,idINfinder,
                                                    linkINfinder,address,roomCount,typeObj,metro,square,floor,maxFloor,
                                                    parking,phone1,phone2,phone3,about,remont,roomSquare,balcon,
                                                    windows,sanuzel,withChildrensPets,additional,height,lift,chute,
                                                    agentComission,buyerComission,buildingType,cost,zalog,squareLive,
                                                    squareKitchen,metroMinutes))
                                self.cianIDs += (idINfinder,)
                            self.con.commit()
            if file.startswith('avito') and file.endswith('.xlsx'):
                # Авито
                wb = openpyxl.load_workbook(filename='./clients/' + self.cmbFolders.currentText() + '/' + file,
                                            read_only=True)
                ws = wb[wb.sheetnames[0]]
                xlsx_header = {}
                for i, row in enumerate(ws):
                    if not i:
                        for j, cell in enumerate(row):
                            xlsx_header[cell.value] = j
                finderType = 2
                linkINfinder, address, metro = '', '', ''
                floor, maxFloor, roomCount, agentComission, buyerComission, idINfinder, cost = 0, 0, 0, 0, 0, 0, 0
                metroMeters = 0
                square = 0.0
                for i, row in enumerate(ws):
                    if i:
                        for j, cell in enumerate(row):
                            if j == xlsx_header['idINfinder']:
                                idINfinder = int(cell.value)
                                if not idINfinder:
                                    continue
                            elif j == xlsx_header['linkINfinder']:
                                linkINfinder = cell.value
                            elif j == xlsx_header['address']:
                                address = cell.value
                            elif j == xlsx_header['metro']:
                                metro = cell.value
                            elif j == xlsx_header['metroMeters']:
                                metroMeters = cell.value
                            elif j == xlsx_header['floor']:
                                floor = int(cell.value)
                            elif j == xlsx_header['maxFloor']:
                                maxFloor = int(cell.value)
                            elif j == xlsx_header['roomCount']:
                                roomCount = int(cell.value)
                            elif j == xlsx_header['agentComission']:
                                agentComission = int(cell.value)
                            elif j == xlsx_header['buyerComission']:
                                buyerComission = int(cell.value)
                            elif j == xlsx_header['square']:
                                square = float(cell.value)
                            elif j == xlsx_header['cost']:
                                cost = float(cell.value)
                        cursorObj = self.con.cursor()
                        if idINfinder:
                            if idINfinder in self.avitoIDs:
                                cursorObj.execute("""UPDATE cards SET linkINfinder = ?, address = ?, metro = ?,
                                metroMeters = ?, floor = ?, maxFloor = ?, roomCount = ?, agentComission = ?, 
                                buyerComission = ?, square = ?, cost = ? WHERE idINfinder = ?""",
                                                  (linkINfinder, address, metro, metroMeters, floor, maxFloor,
                                                   roomCount, agentComission, buyerComission, square, cost, idINfinder))
                            else:
                                cursorObj.execute("INSERT INTO cards (idINfinder, linkINfinder, address, metro, "
                                                  "metroMeters, floor, maxFloor, roomCount, agentComission, "
                                                  "buyerComission, square, cost, finderType) "
                                                  "VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)", (idINfinder, linkINfinder,
                                                                                  address, metro, metroMeters, floor,
                                                                                  maxFloor, roomCount, agentComission,
                                                                                  buyerComission, square, cost,
                                                                                  finderType))
                            self.con.commit()
                            self.avitoIDs += (idINfinder,)

        self.loadBase()
        return

    def leMetro_changed(self):
        if l(self.leMetroMetersMax.text()) < 0 or l(self.leMetroMinutesMax.text()) < 0:
            return
        self.cutMetroMetersMax = l(self.leMetroMetersMax.text())
        self.cutMetroMinutesMax = l(self.leMetroMinutesMax.text())
        self.load_lwCards()
        self.click_lwCards()
        return

    def leComission_changed(self):
        if int(self.leComission.text()) > 100 and int(self.leComission.text()) < 0:
            return
        self.cutComission = int(self.leComission.text())
        self.load_lwCards()
        self.click_lwCards()
        return

    def click_clbLoad(self):
        tableNum = 0
        if self.lwCards.currentIndex().row() > 0:
            tableNum = self.lwCards.currentIndex().row()
        cardNum = self.table2base[tableNum]
        p = Popen(['firefox', self.cards[cardNum][self.colNames['linkINfinder']]])
        return

    def click_clbUpdate(self):
        tableNum = 0
        if self.lwCards.currentIndex().row() > 0:
            tableNum = self.lwCards.currentIndex().row()
        cardNum = self.table2base[tableNum]
        idCard = self.cards[cardNum][self.colNames['id']]
        cursorObj = self.con.cursor()
        cursorObj.execute("UPDATE cards SET phone1 = ?, phone2 = ?, status = ?, showingAT = ?, note = ? WHERE id = ?",
                          (self.lePhone1.text(), self.lePhone2.text(), self.cmbStatus.currentText(),
                           self.dateTimeEdit.dateTime().toPyDateTime(), self.leNote.text(), idCard))
        self.con.commit()
        self.loadBase()
        return

    def click_lwStatuses(self):
        items = self.lwStatuses.selectedItems()
        x = []
        for i in range(len(items)):
            x.append(str(self.lwStatuses.selectedItems()[i].text()))
        self.cutStatuses = tuple(x)
        self.load_lwCards()
        return

    def leCost_changed(self):
        self.cutCostMin = l(self.leCostMin.text())
        self.cutCostMax = l(self.leCostMax.text())
        self.load_lwCards()
        return

    def click_clbTrash(self):
        # Недоделанный скрипт удаления дублей
        cursorObj = self.con.cursor()
        cursorObj.execute("SELECT id,finderType,idINfinder,status FROM cards ")
        rows = cursorObj.fetchall()
        finderTypes = {}
        idINfinders = {}
        idINfinder2ids = {}
        status2id = {}
        for row in rows:
            finderTypes[row[0]] = row[1]
            idINfinders[row[0]] = row[2]
            status2id[row[0]] = row[3]
            if idINfinder2ids[row[2]].get(row[0], None):
                idINfinder2ids[row[2]].append(row[0])
            else:
                idINfinder2ids[row[0]] = [row[0]]
        for idINfinder2id in idINfinder2ids:
            count = len(idINfinder2ids[idINfinder2id])
            if count > 1:
                for id2 in idINfinder2ids[idINfinder2id]:
                    if status2id[id2] == 1:
                        pass
                    else:
                        pass
